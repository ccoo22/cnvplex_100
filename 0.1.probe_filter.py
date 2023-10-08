#!/home/genesky/software/python/3.9.4/bin/python3
import argparse
import glob
import logging
import math
import multiprocessing
import os
import re
import sys
import warnings

import openpyxl
import pysam

warnings.filterwarnings('ignore')
logging.basicConfig(
    format='[%(asctime)s %(levelname)s] %(message)s',
    stream=sys.stdout,
    level=logging.INFO
    )
log = logging.getLogger(__name__)

def set_and_parse_args():
    """参数解析"""
    parser = argparse.ArgumentParser(
        description="探针过滤", formatter_class=argparse.RawTextHelpFormatter)

    parser.add_argument('--input_dir', '-i', type=str,  default="./probe/chromosome",
                        help="探针所在目录： 例如： ./probe/chromosome  ，目录下必须包含子目录，所有的cnvplex设计的原始探针excel表格都必须在子目录下。常用：./probe/core_gene  ./probe/critical_region ./probe/sub_region  ")
    args = parser.parse_args()

    return args
args = set_and_parse_args()
# 对cnvplex设计的原始探针进行过滤，保留合格的探针

# 基因组文件
genome_fa = '/home/genesky/database_new/ucsc/fasta/hg19/samtools_index/hg19.fa'
blast_db = '/home/genesky/database_new/ucsc/fasta/hg19/blast+_index/hg19.fa'
blastn = '/home/genesky/software/blast+/2.13.0/bin/blastn'

def create_dir(dir):
    if not os.path.exists(dir):
        os.makedirs(dir)
# 按照规定做探针过滤
input_dir = args.input_dir

def read_probe(probe_dir):
    probe_info = {}
    duplicate_position = {}
    for subdir in glob.glob(probe_dir + "/*"):
        subname = os.path.basename(subdir)
        # 记录重复的探针位置，防止被多次使用
        for excel in glob.glob(f'{probe_dir}/{subname}/*.xlsx'):
            log.info(f'读入探针文件：{excel}')
            excel_name = os.path.basename(excel)
            wb = openpyxl.load_workbook(excel)
            ws_pair = wb['探针对信息']
            ws_probe = wb['探针信息']
            # （1）读入探针对信息
            # 表头
            heads = [ws_pair.cell(1, column).value for column in range(
                1, ws_pair.max_column + 1)]
            # 读入内容
            for row in range(2, ws_pair.max_row + 1):
                # 取出一行
                tmps = {}
                for column in range(1, ws_pair.max_column + 1):
                    tmps[heads[column - 1]
                         ] = ws_pair.cell(row, column).value
                # 保存
                if tmps['虚拟'] == 'TRUE' or tmps['虚拟'] == True:
                    continue
                probe_name =  subname + "-" + excel_name + '-' + tmps['探针对名称']
                chrom, start, end = tmps['染色体位置'].split('_')
                strand = tmps['探针方向'].split('/')[1]
                positions = [int(start), int(end)]
                positions.sort()

                # 检查是否有重复，重复的结果不用在保存了
                genome_position = '_'.join([chrom, str(positions[0]), str(positions[1]), strand])
                if genome_position in duplicate_position:
                    print("重复探针：", genome_position)
                    continue
                duplicate_position[genome_position] = 1
                probe_info[probe_name] = {}
                # probe_info[probe_name] = tmps
                probe_info[probe_name]['chrom'] = 'chr' + chrom
                probe_info[probe_name]['start'] = positions[0]
                probe_info[probe_name]['end'] = positions[1]
                probe_info[probe_name]['strand'] = strand
                probe_info[probe_name]['position'] = f"{chrom}:{start}-{end}"
                probe_info[probe_name]['probe_name'] = probe_name
                probe_info[probe_name]['最高同源性'] = round(float(tmps['最高同源性']), 2)
                probe_info[probe_name]['CNV%'] = tmps['CNV%']
                probe_info[probe_name]['SNP（MAF）'] = tmps['SNP（MAF）'] if tmps['SNP（MAF）'] != '' else '.'


            # (2) 读入探针序列信息
            heads = [ws_probe.cell(1, column).value for column in range(
                1, ws_probe.max_column + 1)]
            # 读入内容
            for row in range(2, ws_probe.max_row + 1):
                # 取出一行
                tmps = {}
                for column in range(1, ws_probe.max_column + 1):
                    tmps[heads[column - 1]
                         ] = ws_probe.cell(row, column).value
                # 保存
                if tmps['序列信息'] == '' or tmps['序列信息'] == None:
                    continue
                probe_name = subname + "-" + excel_name + '-' + tmps['对应探针对名称']
                # 可能因为坐标重复，而被删掉了
                if probe_name not in probe_info:
                    continue
                
                # 确定是 5‘ 3’
                probe_type_53 = re.match('\d', tmps['探针名称'].split('-')[-1]).group()  # 引物端号： 5 、 3
                probe_info[probe_name][f'{probe_type_53}_length'] = int(tmps['长度'])
                probe_info[probe_name][f'{probe_type_53}_tm'] = round(float(tmps['Tm']), 2)
                if probe_type_53 == '5':
                    probe_info[probe_name][f'5_seq'] = tmps['序列信息'].split(' ')[-1]
                if probe_type_53 == '3':
                    probe_info[probe_name][f'3_seq'] = tmps['序列信息'].split(' ')[0]
            wb.close()
    return probe_info

def apply_6AT(probe_info, output_file):
    log.info(f'6AT 判断')
    # 要求探针 5 3 序列连接点两侧6bp内，必须包含一个A 或 T
    with open(output_file, 'w') as fh:
        for probe_name in probe_info.keys():
            status = True
            desc = '.'
            seq_5 = probe_info[probe_name][f'5_seq']
            seq_3 = probe_info[probe_name][f'3_seq']
            if not re.search('[AT]', seq_5[-6:]):
                status = False
                desc = '5 端没有AT'
            if not re.search('[AT]', seq_3[:6]):
                status = False
                desc = '3 端没有AT'
            fh.write(probe_name + "\t" + str(status) + "\t" + desc + "\n")

def apply_str(probe_info, output_file):
    log.info(f'STR 判断')
    with open(output_file, 'w') as fh:
        for probe_name in probe_info.keys():
            status = True
            desc = '.'
            seq = probe_info[probe_name][f'5_seq'] + probe_info[probe_name][f'3_seq']

            # 1. 单碱基重复不能超过5
            if status:
                match_1 = re.search("(([ATCG]{1})\\2{5,})", seq)
                if match_1:
                    status = False
                    desc = match_1.group(2) + str(len(match_1.group()) / len(match_1.group(2)))
            
            # 2. 2/3 碱基重复，不能大于3
            if status:
                for motif_length in [2, 3]:
                    match23 = re.search("(([ATCG]{%d})\\2{3,})" % (motif_length), seq)
                    if match23:
                        status = False
                        desc = match23.group(2) + str(len(match23.group()) / len(match23.group(2)))
                        break
            # 3. 4 碱基重复，不能大于2
            if status:
                match4 = re.search("(([ATCG]{%d})\\2{2,})" % (4), seq)
                if match4:
                    status = False
                    desc = match4.group(2) + str(len(match4.group()) / len(match4.group(2)))
            # 4. 5以上碱基重复，不能大于1
            if status:
                for motif_length in range(5,20):
                    match_more = re.search("(([ATCG]{%d})\\2{2,})" % (motif_length), seq)
                    if match_more:
                        status = False
                        desc = match_more.group(2) + str(len(match_more.group()) / len(match_more.group(2)))
                        break
            fh.write(probe_name + "\t" + str(status) + "\t" + desc + "\n")

def apply_hom(probe_info, output_file):
    log.info(f'同源性 判断')
    fasta = output_file + ".fa"
    with open(fasta, 'w') as fh:
        for probe_name in probe_info.keys():
            seq = probe_info[probe_name][f'5_seq'] + probe_info[probe_name][f'3_seq']
            fh.write(">"+probe_name + "\n" + seq + "\n")
    # blastn
    blast_res = fasta + ".blastn"
    os.system(f"{blastn} -task blastn -query {fasta} -outfmt 6 -evalue 0.00001 -db {blast_db} -out {blast_res} -num_threads 20 -max_target_seqs 5")
    blast_status = {}
    with open(blast_res, 'r') as fh:
        for line in fh:
            values = line.strip().split()
            probe_name, align_len, chrom = values[0], int(values[3]), values[1]
            if align_len <= 30:  # 比对长度要超过30才行
                continue
            if probe_name not in blast_status:
                blast_status[probe_name] = {'all': 0}
            blast_status[probe_name]['all'] += 1
            blast_status[probe_name][chrom] = blast_status[probe_name].get(chrom, 0) + 1
    # 输出
    with open(output_file, 'w') as fh:
        for probe_name in probe_info.keys():
            blast_num = 1
            if probe_name  in blast_status:
                blast_num = blast_status[probe_name].get('all', 1)
                # 几个特殊染色体探针做特殊处理
                if re.search('chrX-AP0', probe_name) or re.search('chrX-AP1', probe_name) or re.search('chrX-AP2', probe_name) or re.search('chrX-AQ0', probe_name):
                    blast_num = blast_status[probe_name].get('chrX', 0)
                if re.search('chrY-AQ0', probe_name) or re.search('chrY-AQ1', probe_name) or re.search('chrY-AQ2', probe_name) or re.search('chrY-AQ3', probe_name) or re.search('chrY-AQ5', probe_name):
                    blast_num = blast_status[probe_name].get('chrY', 0)
                
            status = True if blast_num == 1 else False
            desc = '.'
            if not status:
                desc = f"blastn {blast_num} position"
            fh.write(probe_name + "\t" + str(status) + "\t" + desc + "\n")


def apply_gc(probe_info, output_file):
    log.info(f'GC 含量判断')
    # 200bp序列的GC含量 35%~65%
    fasta = pysam.FastaFile(genome_fa)
    with open(output_file, 'w') as fh:
        for probe_name in probe_info.keys():
            chrom = probe_info[probe_name]['chrom']
            start = probe_info[probe_name]['start']
            end = probe_info[probe_name]['end']
            probe_len = end - start + 1
            sub = (200 - probe_len) / 2
            # 上部延伸
            start = start - math.ceil(sub)
            if start < 1:  # 防止左侧超界，例如 chr17:56-99
                start = 1
            # 下部延伸
            end = end + math.floor(sub)
            # 取200bp序列
            try:
                seq200 = fasta.fetch(chrom, start-1, end).upper()
            except Exception as err:  # 其他未知的错误，统一打包放在这里
                print("未知错误 %s" % err)
                print(probe_name, chrom, start, end)
            gc = (seq200.count('G') + seq200.count('C')) / 200
            status = True
            desc = 'GC200 ' + str(round(gc, 2))
            if gc < 0.35 or gc > 0.65:
                status = False
            fh.write(probe_name + "\t" + str(status) + "\t" + desc + "\n")

def output(probe_info, status_file_info, final_result):
    log.info(f'汇总输出结果')
    # 读入状态信息
    for name in status_file_info.keys():
        print(status_file_info[name])
        with open(status_file_info[name], 'r') as fh:
            for line in fh:
                probe_name, status, desc = line.strip().split('\t')
                probe_info[probe_name][name] = status
                probe_info[probe_name][name + " desc"] = desc

    # 状态判断
    for probe_name in probe_info.keys():
        good = True
        for name in status_file_info.keys():
            if probe_info[probe_name][name] == 'False':
                good = False
        # snp 判断
        if probe_info[probe_name]['SNP（MAF）'] != '.':
            good = False
        probe_info[probe_name]['check'] = 'PASS' if good else 'ERROR'

    # 输出
    fh_all = open(final_result + ".all.txt", 'w')
    fh_pass = open(final_result, 'w')
    titles = ['probe_name', 'chrom', 'start', 'end', 'strand', '最高同源性', 'CNV%', 'SNP（MAF）', '5_seq', '5_length', '5_tm', '3_seq', '3_length', '3_tm', 'check']
    for name in status_file_info.keys():
        titles.append(name)
        titles.append(name + " desc")
    
    fh_all.write("\t".join(titles) + "\n")
    fh_pass.write("\t".join(titles) + "\n")
    total = 0
    passed = 0
    for probe_name in probe_info.keys():
        total += 1
        values = [str(probe_info[probe_name][title]) for title in titles]
        fh_all.write("\t".join(values) + "\n")
        if probe_info[probe_name]['check'] == 'PASS':
            fh_pass.write("\t".join(values) + "\n")
            passed+=1
    log.info(f'共 {total} 个探针: ' + final_result + ".all.txt")
    log.info(f'合格探针有 {passed} 个 :' + final_result)



# 1. 读入探针信息
probe_info = read_probe(input_dir)

# 2. 探针过滤规则应用
status_gc = os.path.join(input_dir, "gc.txt")
apply_gc(probe_info, status_gc)

status_6at = os.path.join(input_dir, "6at.txt")
apply_6AT(probe_info, status_6at)

status_str = os.path.join(input_dir, "str.txt")
apply_str(probe_info, status_str)

status_homology = os.path.join(input_dir, "hom.txt")
apply_hom(probe_info, status_homology)

# 3. 结果输出
final_result = os.path.join(input_dir, "final.txt")
status_file_info = {
    "GC": status_gc,
    "6AT":status_6at,
    "STR": status_str,
    "hom": status_homology
}
output(probe_info, status_file_info, final_result)

print("分析结果：" + final_result)